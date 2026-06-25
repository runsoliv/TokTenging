"""Pure formatting, parsing, and Excel output helpers for bank imports."""

import datetime
from decimal import Decimal, ROUND_HALF_UP
import re
import unicodedata

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd


def format_date_as_text(date_val):
    """Convert date value to text string format (DD.MM.YYYY)."""
    if pd.isna(date_val):
        return ''
    if isinstance(date_val, pd.Timestamp):
        return date_val.strftime('%d.%m.%Y')
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime('%d.%m.%Y')
    if isinstance(date_val, datetime.date):
        return date_val.strftime('%d.%m.%Y')
    parsed = _format_date_input_line(date_val)
    if parsed:
        return parsed
    # Already a string or other type that the parser does not recognize.
    return str(date_val)

def autofit_excel_columns(file_path):
    """Auto-fit column widths in Excel file and format DATE column as text."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    fixed_widths = {
        "STATUS": 14,
        "Cnf": 4,
        "CONFIDENCE": 4,
    }
    max_widths = {
        "TEXT": 42,
    }
    confidence_fills = {
        "red": PatternFill(fill_type="solid", fgColor="FF0000"),
        "orange": PatternFill(fill_type="solid", fgColor="F4B183"),
        "yellow": PatternFill(fill_type="solid", fgColor="FFFF00"),
        "green": PatternFill(fill_type="solid", fgColor="00B050"),
    }

    # Find special column indexes.
    date_col_idx = None
    confidence_col_idx = None
    for idx, cell in enumerate(ws[1], 1):  # First row (headers)
        if cell.value == 'DATE':
            date_col_idx = idx
        if cell.value in {'Cnf', 'CONFIDENCE'}:
            confidence_col_idx = idx

    def confidence_fill(value):
        try:
            confidence = int(float(str(value).strip()))
        except Exception:
            return None
        if confidence < 50:
            return confidence_fills["red"]
        if confidence < 70:
            return confidence_fills["orange"]
        if confidence < 85:
            return confidence_fills["yellow"]
        return confidence_fills["green"]

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        header = str(col[0].value or "")
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            # Set DATE column format to text (@) to preserve text format on copy-paste
            if date_col_idx and col[0].column == date_col_idx:
                cell.number_format = '@'
            if confidence_col_idx and col[0].column == confidence_col_idx and cell.row > 1:
                fill = confidence_fill(cell.value)
                if fill:
                    cell.fill = fill
        if header in fixed_widths:
            width = fixed_widths[header]
        else:
            width = max_length + 2
            if header in max_widths:
                width = min(width, max_widths[header])
        ws.column_dimensions[column].width = width
    wb.save(file_path)

DATE_MONTH_MAP = {
    # English
    "jan": 1, "january": 1,
    "feb": 2, "febr": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
    # Icelandic, normalized without accents.
    "januar": 1,
    "februar": 2,
    "mars": 3,
    "mai": 5,
    "juni": 6,
    "juli": 7,
    "ag": 8, "agu": 8, "agust": 8,
    "okt": 10, "oktober": 10,
    "des": 12, "desember": 12,
}

DATE_LETTER_RE = r"[^\W\d_]+"

DATE_WEEKDAYS = (
    "monday", "mon", "tuesday", "tue", "tues", "wednesday", "wed",
    "thursday", "thu", "thur", "thurs", "friday", "fri", "saturday",
    "sat", "sunday", "sun", "manudagur", "man", "thridjudagur",
    "thr", "thri", "midvikudagur", "mid", "fimmtudagur", "fim", "fostudagur",
    "fos", "laugardagur", "lau", "sunnudagur", "sun",
)

def _strip_accents(text):
    text = str(text).translate(str.maketrans({
        "Þ": "Th", "þ": "th", "Ð": "D", "ð": "d",
        "Æ": "Ae", "æ": "ae", "Ö": "O", "ö": "o",
    }))
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(ch)
    )

def _month_number(month_raw):
    key = _strip_accents(str(month_raw).strip().strip(".").lower())
    key = re.sub(r"[^a-z]+", "", key)
    return DATE_MONTH_MAP.get(key)

def _normalize_date_input_text(text):
    text = str(text).strip().lower()
    text = text.replace("\u00a0", " ").replace("\t", " ")
    text = re.sub(r"[,]+", " ", text)
    text = re.sub(r"\b(of|the|árið)\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()

    accentless = _strip_accents(text)
    weekday_pattern = "|".join(re.escape(day) for day in DATE_WEEKDAYS)
    if weekday_pattern:
        text = re.sub(
            rf"^(?:{weekday_pattern})\.?\s+",
            "",
            accentless,
            flags=re.IGNORECASE,
        ).strip()
    return text

def _year_from_input(year_raw):
    year = int(year_raw)
    if year < 100:
        year += 2000 if year < 70 else 1900
    return year

def _format_date_parts(day_raw, month_raw, year_raw):
    try:
        day = int(day_raw)
        month = int(month_raw)
        year = _year_from_input(year_raw)
        dt = datetime.date(year, month, day)
        return dt.strftime("%d.%m.%Y")
    except Exception:
        return None

def _format_month_name_date(text):
    patterns = (
        # 5 jan 2026, 5.jan.2026, 5th January 2026, 5. janúar 2026
        (rf"^(\d{{1,2}})(?:st|nd|rd|th)?\.?\s*[-./]?\s*({DATE_LETTER_RE})\.?\s*[-./]?\s*(\d{{2,4}})$", "dmy"),
        # Jan 5 2026, January 5th 2026, janúar 5 2026
        (rf"^({DATE_LETTER_RE})\.?\s*[-./]?\s*(\d{{1,2}})(?:st|nd|rd|th)?\.?\s*[-./]?\s*(\d{{2,4}})$", "mdy"),
        # 2026 Jan 5, 2026 janúar 5
        (rf"^(\d{{4}})\s*[-./]?\s*({DATE_LETTER_RE})\.?\s*[-./]?\s*(\d{{1,2}})(?:st|nd|rd|th)?\.?$", "ymd"),
    )
    for pattern, order in patterns:
        match = re.match(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        first, second, third = match.groups()
        if order == "dmy":
            month = _month_number(second)
            return _format_date_parts(first, month, third) if month else None
        if order == "mdy":
            month = _month_number(first)
            return _format_date_parts(second, month, third) if month else None
        month = _month_number(second)
        return _format_date_parts(third, month, first) if month else None
    return None

def _format_numeric_date(text):
    day_month_match = re.match(r"^(\d{1,2})\s*[\s,./-]+\s*(\d{1,2})$", text)
    if day_month_match:
        day, month = day_month_match.groups()
        return _format_date_parts(day, month, "2026")

    ymd_match = re.match(r"^(\d{4})[.\-/ ]+(\d{1,2})[.\-/ ]+(\d{1,2})$", text)
    if ymd_match:
        year, month, day = ymd_match.groups()
        return _format_date_parts(day, month, year)

    dmy_match = re.match(r"^(\d{1,2})[.\-/ ]+(\d{1,2})[.\-/ ]+(\d{2,4})$", text)
    if dmy_match:
        day, month, year = dmy_match.groups()
        return _format_date_parts(day, month, year)

    if not re.fullmatch(r"[\d\s.,\-/]+", text):
        return None

    digits = re.sub(r"\D", "", text)
    if len(digits) == 8:
        ymd = _format_date_parts(digits[6:8], digits[4:6], digits[0:4])
        if ymd and 1900 <= int(digits[0:4]) <= 2100:
            return ymd
        return _format_date_parts(digits[0:2], digits[2:4], digits[4:8])
    if len(digits) == 6:
        return _format_date_parts(digits[0:2], digits[2:4], digits[4:6])
    return None

def _format_date_input_line(val):
    if pd.isna(val):
        return None
    if isinstance(val, (datetime.datetime, datetime.date, pd.Timestamp)):
        return format_date_as_text(val)

    text = str(val).strip()
    if not text:
        return None

    normalized = _normalize_date_input_text(text)
    parsed = _format_month_name_date(normalized)
    if parsed:
        return parsed

    parsed = _format_numeric_date(normalized)
    if parsed:
        return parsed

    if not re.fullmatch(r"[\d\s.,\-/]+", normalized):
        parsed_dt = pd.to_datetime(text, errors='coerce', dayfirst=True)
        if not pd.isna(parsed_dt):
            return parsed_dt.strftime('%d.%m.%Y')
    return None

def _round_half_up_decimal(val):
    try:
        return Decimal(str(val).replace(" ", "").replace(",", "."))
    except Exception:
        return None

def _parse_bank_amount_decimal(val):
    if pd.isna(val):
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, (int, float)):
        try:
            return Decimal(str(val))
        except Exception:
            return None

    text = str(val).strip()
    if not text:
        return None
    text = text.replace(" ", "").replace("\u00a0", "")
    sign = 1
    if text.startswith("(") and text.endswith(")"):
        sign = -1
        text = text[1:-1]
    if text.startswith("-"):
        sign = -1
        text = text[1:]
    elif text.startswith("+"):
        text = text[1:]
    text = re.sub(r"(?i)(?:isk|kr\.?)$", "", text)
    text = re.sub(r"[^0-9,.]", "", text)

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    elif text.count(".") == 1 and len(text.rsplit(".", 1)[1]) == 3:
        text = text.replace(".", "")
    elif text.count(".") > 1:
        text = text.replace(".", "")

    try:
        return Decimal(text) * sign
    except Exception:
        return None

def _bank_amount_abs(val):
    dec = _parse_bank_amount_decimal(val)
    if dec is None:
        return val
    dec = abs(dec)
    if dec == dec.to_integral_value():
        return int(dec)
    return float(dec)

def _bank_amount_sign(val):
    dec = _parse_bank_amount_decimal(val)
    if dec is None:
        return '+'
    return '+' if dec >= 0 else '-'

def _parse_innheimta_amount(val, force_thousands=False):
    if pd.isna(val):
        return 0
    if isinstance(val, (int,)):
        if force_thousands and abs(val) < 1000:
            return int(val) * 1000
        return int(val)
    if isinstance(val, (float, Decimal)):
        try:
            dec = Decimal(str(val))
        except Exception:
            return 0
        if dec == dec.to_integral_value():
            num = int(dec)
            if force_thousands and abs(num) < 1000:
                return num * 1000
            return num
        # No decimals in source files; non-integer means thousands separator was read as decimal.
        return int((dec * 1000).to_integral_value(rounding=ROUND_HALF_UP))
    text = str(val).strip()
    if not text:
        return 0
    text = text.replace(" ", "").replace("\u00a0", "")
    sign = 1
    if text.startswith("(") and text.endswith(")"):
        sign = -1
        text = text[1:-1]
    if text.startswith("-"):
        sign = -1
        text = text[1:]
    elif text.startswith("+"):
        text = text[1:]
    if not text:
        return 0
    # No decimals in these files; dots are thousands separators.
    text = text.replace(".", "")
    text = text.replace(",", "")
    if not text:
        return 0
    try:
        return sign * int(text)
    except Exception:
        try:
            return sign * int(Decimal(text))
        except Exception:
            return 0


def _format_kennitala(val):
    if pd.isna(val):
        return ''
    text = str(val).strip()
    if not text:
        return ''
    text = text.replace(" ", "").replace("-", "")
    text = text.replace(".", "")
    return text



def _parse_icelandic_date(val):
    parsed = _format_date_input_line(val)
    if parsed:
        return parsed
    if pd.isna(val):
        return ''
    text = str(val).strip()
    return text
