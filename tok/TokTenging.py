import tkinter as tk
from tkinter import ttk, filedialog, PhotoImage
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    TKDND_AVAILABLE = True
    TKDND_ERROR = ""
except Exception as exc:
    DND_FILES = None
    TkinterDnD = None
    TKDND_AVAILABLE = False
    TKDND_ERROR = str(exc)
import pandas as pd
import pyautogui
import time
import logging
import pyperclip
import gc
import datetime
import os
import sys
import json
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import openpyxl
from openpyxl.utils import get_column_letter
import winsound
from threading import Timer

action_delay = 0.1
focus_delay = 0.02
paste_delay = 0.0
post_paste_delay = 0.02
start_delay = 3
running = True
watchdog_timer = None
after_id = None
settings = {}
SETTINGS_PATH = os.path.join(os.path.expanduser("~"), "AppData", "Local", "TokTenging", "settings.json")
pending_tok_df = None
pending_tok_file_path = None
current_tok_file_path = None

# Sentinel used by the action queue to mean "use the current global action_delay".
USE_ACTION_DELAY = object()

# =============================================================================
# MODERN UI THEME COLORS
# =============================================================================
BG_DARK = '#141622'        # Deep background
BG_CARD = '#1f2333'        # Card/panel background
BG_INPUT = '#2a2f42'       # Input field background
BORDER = '#2f3548'         # Subtle borders
ACCENT = '#d6223b'         # Red accent (matches logo)
ACCENT_HOVER = '#ea3f58'   # Lighter accent for hover
TEXT_PRIMARY = '#f5f6fb'   # Near-white text
TEXT_SECONDARY = '#b1b6c6' # Muted text
TEXT_SUCCESS = '#4ade80'   # Green for success
TEXT_ERROR = '#f87171'     # Red for errors
TEXT_WARNING = '#fbbf24'   # Yellow/orange for warnings

# Settings layout helpers
SETTINGS_LABEL_WIDTH = 24

# =============================================================================
# DRAG & DROP HELPERS
# =============================================================================

def _parse_drop_data(data: str) -> str:
    """Extract first file path from tkdnd drop data."""
    if not data:
        return ""
    data = data.strip()
    if not data:
        return ""
    # tkdnd uses a Tcl list format: items may be wrapped in braces if they contain spaces.
    items = []
    buf = ""
    in_brace = False
    for ch in data:
        if ch == "{" and not in_brace:
            in_brace = True
            buf = ""
            continue
        if ch == "}" and in_brace:
            in_brace = False
            items.append(buf)
            buf = ""
            continue
        if ch == " " and not in_brace:
            if buf:
                items.append(buf)
                buf = ""
            continue
        buf += ch
    if buf:
        items.append(buf)
    return items[0] if items else ""

def attach_drop_target(widget, on_path):
    """
    Try to make a widget accept file drops. Returns False if tkdnd is unavailable.
    """
    if not TKDND_AVAILABLE:
        return False
    try:
        try:
            widget.drop_target_register(DND_FILES)
        except Exception:
            widget.drop_target_register('DND_Files')
        widget.dnd_bind('<<Drop>>', lambda e: on_path(_parse_drop_data(e.data)))
        return True
    except Exception:
        return False

def create_styled_button(parent, text, command, width=20, accent=True):
    """Create a modern styled button."""
    bg = ACCENT if accent else BG_CARD
    hover_bg = ACCENT_HOVER if accent else BG_INPUT
    border = ACCENT if accent else BORDER
    btn = tk.Button(parent, text=text, command=command,
                    bg=bg, fg=TEXT_PRIMARY,
                    activebackground=hover_bg, activeforeground=TEXT_PRIMARY,
                    font=('Segoe UI Semibold', 10),
                    relief='flat', cursor='hand2',
                    padx=22, pady=11, width=width, bd=0,
                    highlightthickness=1, highlightbackground=border)
    btn.bind('<Enter>', lambda e: btn.configure(bg=hover_bg))
    btn.bind('<Leave>', lambda e: btn.configure(bg=bg))
    return btn

def create_styled_entry(parent, width=35):
    """Create a modern styled entry field."""
    entry = tk.Entry(parent, width=width,
                     bg=BG_INPUT, fg=TEXT_PRIMARY,
                     insertbackground=TEXT_PRIMARY,
                     font=('Segoe UI', 10),
                     relief='flat', bd=0,
                     highlightthickness=1, highlightbackground=BORDER,
                     highlightcolor=ACCENT)
    return entry

def create_styled_label(parent, text, size=10, color=TEXT_PRIMARY, bold=False, bg=None):
    """Create a modern styled label."""
    weight = 'bold' if bold else 'normal'
    if bg is None:
        try:
            bg = parent.cget("bg")
        except Exception:
            bg = BG_DARK
    label = tk.Label(parent, text=text, bg=bg, fg=color,
                     font=('Segoe UI', size, weight))
    return label

def attach_auto_resize_text(text_widget, min_lines=6, max_lines=14):
    """Auto-resize a Text widget based on line count."""
    def _resize(_event=None):
        try:
            content = text_widget.get("1.0", "end-1c")
        except Exception:
            return
        line_count = content.count("\n") + 1 if content else 1
        new_height = max(min_lines, min(max_lines, line_count))
        if int(text_widget.cget("height")) != new_height:
            text_widget.configure(height=new_height)

    def _on_modified(_event=None):
        try:
            text_widget.edit_modified(False)
        except Exception:
            pass
        _resize()

    text_widget.bind("<<Modified>>", _on_modified)
    text_widget.bind("<KeyRelease>", _resize)
    text_widget.bind("<Control-v>", _resize)
    text_widget.bind("<Control-V>", _resize)
    _resize()
    return _resize

def resource_path(relative_path):
    base_path = getattr(sys, "_MEIPASS", os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
    return os.path.join(base_path, relative_path)

def _default_output_dir():
    return os.path.join(os.path.expanduser("~"), "Desktop", "innlestur")

def load_settings():
    defaults = {
        "action_delay": 0.1,
        "focus_delay": 0.05,
        "paste_delay": 0.0,
        "post_paste_delay": 0.05,
        "start_delay": 3,
        "bank_output_dir": _default_output_dir()
    }
    try:
        if os.path.exists(SETTINGS_PATH):
            with open(SETTINGS_PATH, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            defaults.update(data if isinstance(data, dict) else {})
    except Exception:
        pass
    try:
        defaults["action_delay"] = float(defaults["action_delay"])
    except Exception:
        defaults["action_delay"] = 0.1
    try:
        defaults["focus_delay"] = float(defaults.get("focus_delay", 0.02))
    except Exception:
        defaults["focus_delay"] = 0.02
    try:
        defaults["paste_delay"] = float(defaults.get("paste_delay", 0.0))
    except Exception:
        defaults["paste_delay"] = 0.0
    try:
        defaults["post_paste_delay"] = float(defaults.get("post_paste_delay", 0.02))
    except Exception:
        defaults["post_paste_delay"] = 0.02
    try:
        defaults["start_delay"] = int(defaults["start_delay"])
    except Exception:
        defaults["start_delay"] = 3
    if not defaults.get("bank_output_dir"):
        defaults["bank_output_dir"] = _default_output_dir()
    return defaults

def save_settings():
    try:
        os.makedirs(os.path.dirname(SETTINGS_PATH), exist_ok=True)
        with open(SETTINGS_PATH, "w", encoding="utf-8") as handle:
            json.dump(settings, handle, indent=2)
    except Exception:
        pass

def _after_ms(seconds):
    # Tkinter expects milliseconds. Always schedule at least 1ms.
    try:
        return max(1, int(float(seconds) * 1000))
    except Exception:
        return 1

def _cancel_scheduled_after():
    global after_id
    if after_id is not None:
        try:
            root.after_cancel(after_id)
        except Exception:
            pass
        after_id = None

def _cancel_watchdog():
    global watchdog_timer
    if watchdog_timer:
        try:
            watchdog_timer.cancel()
        except Exception:
            pass
    watchdog_timer = None

def _sync_settings_from_ui():
    global action_delay, start_delay, focus_delay, paste_delay, post_paste_delay
    if "settings_action_entry" in globals() and settings_action_entry and settings_action_entry.winfo_exists():
        try:
            action_delay = float(settings_action_entry.get())
        except Exception:
            pass
        settings["action_delay"] = action_delay
    if "settings_focus_entry" in globals() and settings_focus_entry and settings_focus_entry.winfo_exists():
        try:
            focus_delay = float(settings_focus_entry.get())
        except Exception:
            pass
        settings["focus_delay"] = focus_delay
    if "settings_paste_entry" in globals() and settings_paste_entry and settings_paste_entry.winfo_exists():
        try:
            paste_delay = float(settings_paste_entry.get())
        except Exception:
            pass
        settings["paste_delay"] = paste_delay
    if "settings_post_paste_entry" in globals() and settings_post_paste_entry and settings_post_paste_entry.winfo_exists():
        try:
            post_paste_delay = float(settings_post_paste_entry.get())
        except Exception:
            pass
        settings["post_paste_delay"] = post_paste_delay
    if "settings_start_entry" in globals() and settings_start_entry and settings_start_entry.winfo_exists():
        try:
            start_delay = int(float(settings_start_entry.get()))
        except Exception:
            pass
        settings["start_delay"] = start_delay
    if "settings_output_dir_entry" in globals() and settings_output_dir_entry and settings_output_dir_entry.winfo_exists():
        output_dir = settings_output_dir_entry.get().strip()
        if output_dir:
            settings["bank_output_dir"] = output_dir
    save_settings()

def ensure_window_fits():
    root.update_idletasks()
    req_w = root.winfo_reqwidth()
    req_h = root.winfo_reqheight()
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    target_w = min(max(root.winfo_width(), req_w), screen_w - 40)
    target_h = min(max(root.winfo_height(), req_h), screen_h - 60)
    root.geometry(f"{int(target_w)}x{int(target_h)}")

def _stop_automation_only():
    """Stop any currently running auto-input as fast as possible (non-blocking)."""
    global running, actions
    running = False
    actions = []
    _cancel_scheduled_after()
    _cancel_watchdog()

def clear_terminal():
    if os.name == 'nt': 
        os.system('cls')

def stop_script():
    _stop_automation_only()
    logging.info("Stopping...")
    display_stopped_screen()

def display_stopped_screen():
    """Show a summary screen when the script is stopped manually."""
    for widget in frame.winfo_children():
        widget.destroy()
    
    # Calculate runtime
    end_time = datetime.datetime.now()
    time_elapsed = end_time - start_time if 'start_time' in globals() and start_time else datetime.timedelta(0)
    
    # Get current row info
    current_row_num = row_index + 1 if 'row_index' in globals() else 0
    total_rows_count = len(rows) if 'rows' in globals() and rows else 0
    
    # Header with icon
    stopped_label = create_styled_label(frame, "⏹ Script Stopped", size=16, color=TEXT_ERROR, bold=True)
    stopped_label.pack(pady=(20, 15))
    
    # Info card
    card = tk.Frame(frame, bg=BG_CARD, padx=20, pady=15)
    card.pack(fill=tk.X, padx=20, pady=10)
    
    # Row info
    row_info_label = tk.Label(card, text=f"Stopped at Row: {current_row_num} / {total_rows_count}",
                              bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 11, 'bold'))
    row_info_label.pack(pady=5)
    
    # Show current row details if available (truncate long text)
    if 'rows' in globals() and rows and row_index < len(rows):
        current_row = rows[row_index]
        date_val = str(current_row.get('DATE', 'N/A'))
        text_val = str(current_row.get('TEXT', 'N/A'))[:30] + ('...' if len(str(current_row.get('TEXT', ''))) > 30 else '')
        amount_val = str(current_row.get('AMOUNT', 'N/A'))
        
        details_frame = tk.Frame(card, bg=BG_CARD)
        details_frame.pack(pady=10)
        
        for label, value in [("DATE", date_val), ("TEXT", text_val), ("AMOUNT", amount_val)]:
            row_frame = tk.Frame(details_frame, bg=BG_CARD)
            row_frame.pack(fill=tk.X, pady=2)
            tk.Label(row_frame, text=f"{label}:", bg=BG_CARD, fg=TEXT_SECONDARY,
                    font=('Segoe UI', 9), width=8, anchor='e').pack(side=tk.LEFT)
            tk.Label(row_frame, text=value, bg=BG_CARD, fg=TEXT_PRIMARY,
                    font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=5)
    
    # Stats card
    stats_card = tk.Frame(frame, bg=BG_CARD, padx=20, pady=15)
    stats_card.pack(fill=tk.X, padx=20, pady=10)
    
    stats_frame = tk.Frame(stats_card, bg=BG_CARD)
    stats_frame.pack()
    
    # Runtime
    tk.Label(stats_frame, text="⏱ Runtime:", bg=BG_CARD, fg=TEXT_SECONDARY,
             font=('Segoe UI', 10)).pack(side=tk.LEFT)
    tk.Label(stats_frame, text=format_timedelta(time_elapsed), bg=BG_CARD, fg=TEXT_PRIMARY,
             font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT, padx=(5, 20))
    
    # Rows processed
    tk.Label(stats_frame, text="📊 Rows:", bg=BG_CARD, fg=TEXT_SECONDARY,
             font=('Segoe UI', 10)).pack(side=tk.LEFT)
    tk.Label(stats_frame, text=str(row_index if 'row_index' in globals() else 0), bg=BG_CARD, fg=TEXT_PRIMARY,
             font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT, padx=5)
    
    # Back button
    if current_tok_file_path:
        open_btn = create_styled_button(frame, "Open File", lambda: os.startfile(current_tok_file_path), width=20, accent=False)
        open_btn.pack(pady=(5, 10))

    back_button = create_styled_button(frame, "← Back to Menu", initialize_main_menu, width=20)
    back_button.pack(pady=20)

def format_number(number):
    number_str = str(number).split('.')[0]
    return number_str.replace(',', '')

def enter_data(row):
    date_str = row['DATE'].strftime('%Y-%m-%d') if isinstance(row['DATE'], pd.Timestamp) else row['DATE']
    actions = []
    field_delay = 0.03

    def _step(func, delay=USE_ACTION_DELAY):
        return (func, delay)

    def _press_enter():
        pyautogui.press('enter')

    def _paste_text_steps(text):
        # Split into multiple steps to keep the UI responsive and allow Stop mid-row.
        text = "" if text is None else str(text)
        return [
            _step(lambda t=text: pyperclip.copy(t), paste_delay),
            _step(lambda: pyautogui.hotkey('ctrl', 'v'), post_paste_delay),
        ]

    actions.extend(_paste_text_steps(date_str))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(lambda: None, field_delay))

    actions.extend(_paste_text_steps(str(row['TEXT'])))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(lambda: None, field_delay))

    actions.extend(_paste_text_steps(format_number(row['DEBIT'])))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(lambda: None, field_delay))

    if pd.notna(row['ID']):
        actions.extend(_paste_text_steps(format_number(row['ID'])))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(lambda: None, field_delay))

    actions.extend(_paste_text_steps(format_number(row['AMOUNT'])))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(lambda: None, field_delay))
    actions.extend(_paste_text_steps(format_number(row['CREDIT'])))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(lambda: None, field_delay))

    if pd.notna(row['ID']):
        actions.extend(_paste_text_steps(format_number(row['ID'])))
    actions.append(_step(_press_enter, focus_delay))
    actions.append(_step(_press_enter, focus_delay))

    return actions

def write_text_via_clipboard(text):
    """
    TODO(deprecated): Kept for compatibility, but the auto-input now uses a non-blocking
    step queue (see `enter_data`) so Stop can interrupt mid-row.
    """
    pyperclip.copy("" if text is None else str(text))
    pyautogui.hotkey('ctrl', 'v')

def init_progress_bar(total_rows):
    for widget in frame.winfo_children():
        widget.destroy()

    global progress_bar, progress_label, stop_button_processing

    # Title
    title_label = create_styled_label(frame, "Processing...", size=14, bold=True)
    title_label.pack(pady=(20, 15))

    progress_label = create_styled_label(frame, "", size=10, color=TEXT_SECONDARY)
    progress_label.pack(pady=10)

    # Progress bar with custom style
    progress_bar = ttk.Progressbar(frame, orient='horizontal', length=350, mode='determinate',
                                   style='Horizontal.TProgressbar')
    progress_bar.pack(pady=15)
    progress_bar['maximum'] = total_rows
    progress_bar['value'] = 0
    
    stop_button_processing = create_styled_button(frame, "⏹ Stop", stop_script, width=15, accent=False)
    stop_button_processing.pack(pady=20)

def update_progress(row_count):
    progress_bar['value'] = row_count
    progress_label.config(text=f"Processing... {row_count}/{progress_bar['maximum']} rows")
    root.update()

def display_results(time_elapsed, rows_processed):
    for widget in frame.winfo_children():
        widget.destroy()

    # Success header
    success_label = create_styled_label(frame, "✓ Completed!", size=18, color=TEXT_SUCCESS, bold=True)
    success_label.pack(pady=(30, 20))
    
    # Stats card
    card = tk.Frame(frame, bg=BG_CARD, padx=30, pady=20)
    card.pack(padx=20, pady=10)
    
    tk.Label(card, text=f"⏱ Time: {format_timedelta(time_elapsed)}", bg=BG_CARD, fg=TEXT_PRIMARY,
             font=('Segoe UI', 11)).pack(pady=5)
    tk.Label(card, text=f"📊 Rows Processed: {rows_processed}", bg=BG_CARD, fg=TEXT_PRIMARY,
             font=('Segoe UI', 11)).pack(pady=5)
    
    play_success_sound()
    
    back_button = create_styled_button(frame, "← Back to Menu", initialize_main_menu, width=20)
    back_button.pack(pady=25)

def format_timedelta(td):
    minutes, seconds = divmod(td.total_seconds(), 60)
    hours, minutes = divmod(minutes, 60)
    return f"{int(hours):02}:{int(minutes):02}:{seconds:.2f}"

def display_input_controls():
    global file_path_entry, browse_button, run_button
    global input_controls_frame

    # Container for input controls
    input_controls_frame = tk.Frame(frame, bg=BG_CARD, highlightthickness=1, highlightbackground=BORDER)
    input_controls_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
    inner_controls = tk.Frame(input_controls_frame, bg=BG_CARD)
    inner_controls.pack(fill=tk.X, padx=16, pady=14)

    # File selection row with drop area
    file_frame = tk.Frame(inner_controls, bg=BG_CARD)
    file_frame.pack(fill=tk.X, pady=6)
    
    drop_area = tk.Label(file_frame, text="Drop file here", bg=BG_INPUT, fg=TEXT_SECONDARY,
                         font=('Segoe UI', 10), width=24, height=5, relief='flat',
                         highlightthickness=1, highlightbackground=BORDER)
    drop_area.pack(side=tk.LEFT, padx=(0, 12), pady=6, fill=tk.X, expand=True)
    
    file_path_entry = create_styled_entry(file_frame, width=35)
    file_path_entry.pack_forget()  # hidden; we'll set it from drops/browse
    
    def on_drop_tok(path):
        if path:
            file_path_entry.delete(0, tk.END)
            file_path_entry.insert(0, path)
            drop_area.config(text=os.path.basename(path), fg=TEXT_PRIMARY)
    if not attach_drop_target(drop_area, on_drop_tok):
        drop_area.config(text="Install tkinterdnd2 for drag-and-drop", fg=TEXT_WARNING)
    
    def browse_tok():
        filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select a File",
                                              filetypes=(("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")))
        if filename:
            on_drop_tok(filename)
    browse_button = create_styled_button(file_frame, "Browse", browse_tok, width=12, accent=False)
    browse_button.pack(side=tk.LEFT)

    # Run button
    btn_frame = tk.Frame(inner_controls, bg=BG_CARD)
    btn_frame.pack(fill=tk.X, pady=(12, 4))
    
    run_button = create_styled_button(btn_frame, "▶ Run", run_script_from_gui, width=20)
    run_button.pack()

def run_script_with_df(df):
    global running, actions, rows, row_index, start_time, watchdog_timer
    running = True
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')

    total_rows = len(df)
    start_time = datetime.datetime.now()
    rows = df.to_dict(orient='records')
    actions = enter_data(rows[0])
    row_index = 0
    _cancel_scheduled_after()
    _cancel_watchdog()
    # Non-blocking delay before starting automation.
    root.after(2000, process_next_action)

def display_tok_input_error(message):
    for widget in frame.winfo_children():
        widget.destroy()
    title_label = create_styled_label(frame, "Incompatible File", size=14, color=TEXT_ERROR, bold=True)
    title_label.pack(pady=(20, 10))
    msg_label = create_styled_label(frame, message, size=10, color=TEXT_SECONDARY)
    msg_label.pack(pady=(0, 15))
    back_button = create_styled_button(frame, "← Back to Menu", initialize_main_menu, width=20)
    back_button.pack(pady=10)
    ensure_window_fits()

def display_tok_missing_fields_prompt(missing_rows):
    for widget in frame.winfo_children():
        widget.destroy()
    title_label = create_styled_label(frame, "Missing Debit/Credit", size=14, color=TEXT_WARNING, bold=True)
    title_label.pack(pady=(20, 10))

    if len(missing_rows) > 10:
        shown = ", ".join(str(r) for r in missing_rows[:10]) + f" (+{len(missing_rows) - 10} more)"
    else:
        shown = ", ".join(str(r) for r in missing_rows)
    msg_label = create_styled_label(frame, f"Missing DEBIT/CREDIT on row(s): {shown}", size=10, color=TEXT_SECONDARY)
    msg_label.pack(pady=(0, 15))

    def continue_anyway():
        if pending_tok_df is not None and pending_tok_file_path:
            start_tok_run(pending_tok_df, pending_tok_file_path)

    def open_file():
        if pending_tok_file_path:
            os.startfile(pending_tok_file_path)

    btn_frame = tk.Frame(frame, bg=BG_DARK)
    btn_frame.pack(pady=5)
    continue_btn = create_styled_button(btn_frame, "Continue Anyway", continue_anyway, width=20)
    continue_btn.pack(pady=5)
    open_btn = create_styled_button(btn_frame, "Open File", open_file, width=20, accent=False)
    open_btn.pack(pady=5)

    back_button = create_styled_button(frame, "← Back to Menu", initialize_main_menu, width=20)
    back_button.pack(pady=10)
    ensure_window_fits()

def start_tok_run(df, file_path):
    global pending_tok_df, pending_tok_file_path, current_tok_file_path
    pending_tok_df = df
    pending_tok_file_path = file_path
    current_tok_file_path = file_path
    total_rows = len(df)
    init_progress_bar(total_rows)
    countdown_label = create_styled_label(frame, f"Starting in {int(start_delay)} seconds...", size=11, color=TEXT_WARNING)
    countdown_label.pack(pady=10)

    def update_countdown(seconds_left):
        countdown_label.config(text=f"Starting in {seconds_left} seconds...")
        # Keep progress bar at zero during countdown
        progress_bar['value'] = 0
        progress_label.config(text=f"Rows processed: 0 / {total_rows}")
        if seconds_left > 0:
            root.after(1000, update_countdown, seconds_left - 1)
        else:
            countdown_label.pack_forget()
            # Non-blocking: keep UI responsive so Stop works immediately.
            root.after(1000, lambda: run_script_with_df(df))

    update_countdown(int(start_delay))

def process_next_action():
    global row_index, actions, rows, watchdog_timer, after_id
    if not running:
        return

    _cancel_watchdog()

    if actions:
        func, delay = actions.pop(0)
        if not running:
            return
        try:
            func()
        except Exception:
            logging.exception("Action failed; stopping automation.")
            _stop_automation_only()
            return
        watchdog_timer = Timer(15, on_watchdog_timeout)
        watchdog_timer.start()
        next_delay = action_delay if delay is USE_ACTION_DELAY else delay
        after_id = root.after(_after_ms(next_delay), process_next_action)
    else:
        row_index += 1
        if row_index >= len(rows) or rows[row_index]['DATE'] == 'xx':
            complete_script()
        else:
            actions = enter_data(rows[row_index])
            update_progress(row_index)
            watchdog_timer = Timer(15, on_watchdog_timeout)
            watchdog_timer.start()
            after_id = root.after(_after_ms(action_delay), process_next_action)

def complete_script():
    global start_time, watchdog_timer
    _cancel_watchdog()
    end_time = datetime.datetime.now()
    time_elapsed = end_time - start_time
    display_results(time_elapsed, row_index)
    gc.collect()
    logging.info("Completed.")
    logging.info(f"Total time elapsed: {time_elapsed.total_seconds():.2f} seconds")
    logging.info(f"Number of rows processed: {row_index}")

def run_script_from_gui():
    file_path = file_path_entry.get()
    global current_tok_file_path
    current_tok_file_path = file_path
    _sync_settings_from_ui()
    try:
        df = pd.read_excel(file_path)
    except Exception as exc:
        display_tok_input_error(f"Failed to read file: {str(exc)[:80]}")
        return

    df = df.dropna(how='all')

    required_cols = ['DATE', 'TEXT', 'DEBIT', 'ID', 'AMOUNT', 'CREDIT']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        display_tok_input_error(f"Missing column(s): {', '.join(missing_cols)}")
        return

    def _parse_decimal(value):
        if pd.isna(value):
            return None
        if isinstance(value, str):
            cleaned = value.replace(" ", "").replace(",", ".")
        else:
            cleaned = str(value)
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None

    has_decimals = False
    for val in df['AMOUNT']:
        dec = _parse_decimal(val)
        if dec is not None and dec != dec.to_integral_value():
            has_decimals = True
            break
    if has_decimals:
        def _round_half_up(val):
            dec = _parse_decimal(val)
            if dec is None:
                return val
            return int(dec.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        df['AMOUNT'] = df['AMOUNT'].apply(_round_half_up)

    df['DATE'] = df['DATE'].apply(format_date_as_text)

    def _is_missing(val):
        if pd.isna(val):
            return True
        if isinstance(val, str) and not val.strip():
            return True
        return False

    missing_rows = []
    for idx, row in df.iterrows():
        has_data = False
        for col in ['DATE', 'TEXT', 'ID', 'AMOUNT']:
            val = row.get(col)
            if pd.isna(val):
                continue
            if isinstance(val, str) and not val.strip():
                continue
            has_data = True
            break
        if has_data and (_is_missing(row.get('DEBIT')) or _is_missing(row.get('CREDIT'))):
            missing_rows.append(idx + 2)  # Excel row number (header is row 1)

    if missing_rows:
        global pending_tok_df, pending_tok_file_path
        pending_tok_df = df
        pending_tok_file_path = file_path
        display_tok_missing_fields_prompt(missing_rows)
        return

    start_tok_run(df, file_path)

def browse_files():
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select a File", filetypes=(("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")))
    file_path_entry.delete(0, tk.END)
    file_path_entry.insert(0, filename)

def increase_start_delay():
    current_value = int(settings_start_entry.get())
    settings_start_entry.delete(0, tk.END)
    settings_start_entry.insert(0, str(current_value + 1))
    _sync_settings_from_ui()

def decrease_start_delay():
    current_value = int(settings_start_entry.get())
    current_value = max(0, current_value - 1)
    settings_start_entry.delete(0, tk.END)
    settings_start_entry.insert(0, str(current_value))
    _sync_settings_from_ui()

def increase_action_delay():
    global action_delay
    current_value = float(settings_action_entry.get())
    new_value = current_value + 0.001
    settings_action_entry.delete(0, tk.END)
    settings_action_entry.insert(0, f"{new_value:.3f}")
    action_delay = new_value
    _sync_settings_from_ui()

def decrease_action_delay():
    global action_delay
    current_value = float(settings_action_entry.get())
    new_value = max(0, current_value - 0.001)
    settings_action_entry.delete(0, tk.END)
    settings_action_entry.insert(0, f"{new_value:.3f}")
    action_delay = new_value
    _sync_settings_from_ui()

def increase_focus_delay():
    global focus_delay
    current_value = float(settings_focus_entry.get())
    new_value = current_value + 0.001
    settings_focus_entry.delete(0, tk.END)
    settings_focus_entry.insert(0, f"{new_value:.3f}")
    focus_delay = new_value
    _sync_settings_from_ui()

def decrease_focus_delay():
    global focus_delay
    current_value = float(settings_focus_entry.get())
    new_value = max(0, current_value - 0.001)
    settings_focus_entry.delete(0, tk.END)
    settings_focus_entry.insert(0, f"{new_value:.3f}")
    focus_delay = new_value
    _sync_settings_from_ui()

def increase_paste_delay():
    global paste_delay
    current_value = float(settings_paste_entry.get())
    new_value = current_value + 0.001
    settings_paste_entry.delete(0, tk.END)
    settings_paste_entry.insert(0, f"{new_value:.3f}")
    paste_delay = new_value
    _sync_settings_from_ui()

def decrease_paste_delay():
    global paste_delay
    current_value = float(settings_paste_entry.get())
    new_value = max(0, current_value - 0.001)
    settings_paste_entry.delete(0, tk.END)
    settings_paste_entry.insert(0, f"{new_value:.3f}")
    paste_delay = new_value
    _sync_settings_from_ui()

def increase_post_paste_delay():
    global post_paste_delay
    current_value = float(settings_post_paste_entry.get())
    new_value = current_value + 0.001
    settings_post_paste_entry.delete(0, tk.END)
    settings_post_paste_entry.insert(0, f"{new_value:.3f}")
    post_paste_delay = new_value
    _sync_settings_from_ui()

def decrease_post_paste_delay():
    global post_paste_delay
    current_value = float(settings_post_paste_entry.get())
    new_value = max(0, current_value - 0.001)
    settings_post_paste_entry.delete(0, tk.END)
    settings_post_paste_entry.insert(0, f"{new_value:.3f}")
    post_paste_delay = new_value
    _sync_settings_from_ui()

def initialize_main_menu():
    _sync_settings_from_ui()
    for widget in frame.winfo_children():
        widget.destroy()
    
    # Title
    title_label = create_styled_label(frame, "Veldu aðgerð", size=14, bold=True)
    title_label.pack(pady=(20, 30))

    # Menu card
    menu_card = tk.Frame(frame, bg=BG_CARD, highlightthickness=1, highlightbackground=BORDER)
    menu_card.pack(fill=tk.X, padx=20, pady=10)
    menu_inner = tk.Frame(menu_card, bg=BG_CARD)
    menu_inner.pack(fill=tk.X, padx=16, pady=14)

    # Menu buttons
    tok_input_button = create_styled_button(menu_inner, "📝  Tok Input", initialize_tok_input, width=25)
    tok_input_button.pack(pady=8)
    
    bank_formatter_button = create_styled_button(menu_inner, "🏦  Bank Formatter", initialize_bank_formatter, width=25)
    bank_formatter_button.pack(pady=8)

    round_numbers_button = create_styled_button(menu_inner, "🔢  Round Numbers", initialize_round_numbers, width=25)
    round_numbers_button.pack(pady=8)

    format_dates_button = create_styled_button(menu_inner, "📅  Format Dates", initialize_format_dates, width=25)
    format_dates_button.pack(pady=8)

    id_button = create_styled_button(menu_inner, "🆔  Format ID Numbers", initialize_format_ids, width=25)
    id_button.pack(pady=8)

    settings_button = create_styled_button(menu_inner, "⚙  Settings", initialize_settings, width=25, accent=True)
    settings_button.pack(pady=8)

def initialize_tok_input():
    for widget in frame.winfo_children():
        widget.destroy()
    
    title_label = create_styled_label(frame, "Tok Input", size=14, bold=True)
    title_label.pack(pady=(10, 20))
    
    display_input_controls()
    
    back_button = create_styled_button(frame, "← Back to Menu", initialize_main_menu, width=25, accent=False)
    back_button.pack(pady=20)
    ensure_window_fits()

def initialize_settings():
    for widget in frame.winfo_children():
        widget.destroy()

    title_label = create_styled_label(frame, "Settings", size=14, bold=True)
    title_label.pack(pady=(10, 20))

    settings_card = tk.Frame(frame, bg=BG_CARD, highlightthickness=1, highlightbackground=BORDER)
    settings_card.pack(fill=tk.X, padx=20, pady=(0, 10))
    settings_frame = tk.Frame(settings_card, bg=BG_CARD)
    settings_frame.pack(fill=tk.X, padx=16, pady=14)

    # Action delay row
    action_frame = tk.Frame(settings_frame, bg=BG_CARD)
    action_frame.pack(fill=tk.X, pady=10)
    action_label = create_styled_label(action_frame, "Action Delay (sec):", size=9, color=TEXT_SECONDARY)
    action_label.configure(width=SETTINGS_LABEL_WIDTH, anchor='w')
    action_label.pack(side=tk.LEFT)

    global settings_action_entry
    settings_action_entry = tk.Entry(action_frame, width=8, bg=BG_INPUT, fg=TEXT_PRIMARY,
                                     insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                                     relief='flat', justify='center', bd=0,
                                     highlightthickness=1, highlightbackground=BORDER,
                                     highlightcolor=ACCENT)
    settings_action_entry.insert(0, f"{settings.get('action_delay', 0.1):.3f}")
    settings_action_entry.pack(side=tk.LEFT, padx=(6, 12))

    action_dec = tk.Button(action_frame, text="−", command=decrease_action_delay,
                           bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 12, 'bold'),
                           relief='flat', width=3, cursor='hand2',
                           highlightthickness=1, highlightbackground=BORDER)
    action_dec.pack(side=tk.LEFT, padx=(0, 6))

    action_inc = tk.Button(action_frame, text="+", command=increase_action_delay,
                           bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 12, 'bold'),
                           relief='flat', width=3, cursor='hand2',
                           highlightthickness=1, highlightbackground=BORDER)
    action_inc.pack(side=tk.LEFT, padx=(0, 0))

    # Focus delay row
    focus_frame = tk.Frame(settings_frame, bg=BG_CARD)
    focus_frame.pack(fill=tk.X, pady=10)
    focus_label = create_styled_label(focus_frame, "Focus delay (sec):", size=9, color=TEXT_SECONDARY)
    focus_label.configure(width=SETTINGS_LABEL_WIDTH, anchor='w')
    focus_label.pack(side=tk.LEFT)

    global settings_focus_entry
    settings_focus_entry = tk.Entry(focus_frame, width=8, bg=BG_INPUT, fg=TEXT_PRIMARY,
                                    insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                                    relief='flat', justify='center', bd=0,
                                    highlightthickness=1, highlightbackground=BORDER,
                                    highlightcolor=ACCENT)
    settings_focus_entry.insert(0, f"{settings.get('focus_delay', 0.02):.3f}")
    settings_focus_entry.pack(side=tk.LEFT, padx=(6, 12))

    focus_dec = tk.Button(focus_frame, text="−", command=decrease_focus_delay,
                          bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 12, 'bold'),
                          relief='flat', width=3, cursor='hand2',
                          highlightthickness=1, highlightbackground=BORDER)
    focus_dec.pack(side=tk.LEFT, padx=(0, 6))

    focus_inc = tk.Button(focus_frame, text="+", command=increase_focus_delay,
                          bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 12, 'bold'),
                          relief='flat', width=3, cursor='hand2',
                          highlightthickness=1, highlightbackground=BORDER)
    focus_inc.pack(side=tk.LEFT, padx=(0, 0))

    # Paste delay row
    paste_frame = tk.Frame(settings_frame, bg=BG_CARD)
    paste_frame.pack(fill=tk.X, pady=10)
    paste_label = create_styled_label(paste_frame, "Paste delay (sec):", size=9, color=TEXT_SECONDARY)
    paste_label.configure(width=SETTINGS_LABEL_WIDTH, anchor='w')
    paste_label.pack(side=tk.LEFT)

    global settings_paste_entry
    settings_paste_entry = tk.Entry(paste_frame, width=8, bg=BG_INPUT, fg=TEXT_PRIMARY,
                                    insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                                    relief='flat', justify='center', bd=0,
                                    highlightthickness=1, highlightbackground=BORDER,
                                    highlightcolor=ACCENT)
    settings_paste_entry.insert(0, f"{settings.get('paste_delay', 0.0):.3f}")
    settings_paste_entry.pack(side=tk.LEFT, padx=(6, 12))

    paste_dec = tk.Button(paste_frame, text="−", command=decrease_paste_delay,
                          bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 12, 'bold'),
                          relief='flat', width=3, cursor='hand2',
                          highlightthickness=1, highlightbackground=BORDER)
    paste_dec.pack(side=tk.LEFT, padx=(0, 6))

    paste_inc = tk.Button(paste_frame, text="+", command=increase_paste_delay,
                          bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 12, 'bold'),
                          relief='flat', width=3, cursor='hand2',
                          highlightthickness=1, highlightbackground=BORDER)
    paste_inc.pack(side=tk.LEFT, padx=(0, 0))

    # Post-paste delay row
    post_frame = tk.Frame(settings_frame, bg=BG_CARD)
    post_frame.pack(fill=tk.X, pady=10)
    post_label = create_styled_label(post_frame, "Post-paste delay (sec):", size=9, color=TEXT_SECONDARY)
    post_label.configure(width=SETTINGS_LABEL_WIDTH, anchor='w')
    post_label.pack(side=tk.LEFT)

    global settings_post_paste_entry
    settings_post_paste_entry = tk.Entry(post_frame, width=8, bg=BG_INPUT, fg=TEXT_PRIMARY,
                                         insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                                         relief='flat', justify='center', bd=0,
                                         highlightthickness=1, highlightbackground=BORDER,
                                         highlightcolor=ACCENT)
    settings_post_paste_entry.insert(0, f"{settings.get('post_paste_delay', 0.02):.3f}")
    settings_post_paste_entry.pack(side=tk.LEFT, padx=(6, 12))

    post_dec = tk.Button(post_frame, text="−", command=decrease_post_paste_delay,
                         bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 12, 'bold'),
                         relief='flat', width=3, cursor='hand2',
                         highlightthickness=1, highlightbackground=BORDER)
    post_dec.pack(side=tk.LEFT, padx=(0, 6))

    post_inc = tk.Button(post_frame, text="+", command=increase_post_paste_delay,
                         bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 12, 'bold'),
                         relief='flat', width=3, cursor='hand2',
                         highlightthickness=1, highlightbackground=BORDER)
    post_inc.pack(side=tk.LEFT, padx=(0, 0))

    # Start delay row
    start_frame = tk.Frame(settings_frame, bg=BG_CARD)
    start_frame.pack(fill=tk.X, pady=10)
    start_label = create_styled_label(start_frame, "Start Delay (sec):", size=9, color=TEXT_SECONDARY)
    start_label.configure(width=SETTINGS_LABEL_WIDTH, anchor='w')
    start_label.pack(side=tk.LEFT)

    global settings_start_entry
    settings_start_entry = tk.Entry(start_frame, width=8, bg=BG_INPUT, fg=TEXT_PRIMARY,
                                    insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                                    relief='flat', justify='center', bd=0,
                                    highlightthickness=1, highlightbackground=BORDER,
                                    highlightcolor=ACCENT)
    settings_start_entry.insert(0, str(settings.get("start_delay", 3)))
    settings_start_entry.pack(side=tk.LEFT, padx=(6, 12))

    start_dec = tk.Button(start_frame, text="−", command=decrease_start_delay,
                          bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 12, 'bold'),
                          relief='flat', width=3, cursor='hand2',
                          highlightthickness=1, highlightbackground=BORDER)
    start_dec.pack(side=tk.LEFT, padx=(0, 6))

    start_inc = tk.Button(start_frame, text="+", command=increase_start_delay,
                          bg=BG_CARD, fg=TEXT_PRIMARY, font=('Segoe UI', 12, 'bold'),
                          relief='flat', width=3, cursor='hand2',
                          highlightthickness=1, highlightbackground=BORDER)
    start_inc.pack(side=tk.LEFT, padx=(0, 0))

    # Bank output folder
    output_frame = tk.Frame(settings_frame, bg=BG_CARD)
    output_frame.pack(fill=tk.X, pady=10)
    output_label = create_styled_label(output_frame, "Bank output folder:", size=9, color=TEXT_SECONDARY)
    output_label.pack(anchor='w')

    global settings_output_dir_entry
    settings_output_dir_entry = tk.Entry(output_frame, width=42, bg=BG_INPUT, fg=TEXT_PRIMARY,
                                         insertbackground=TEXT_PRIMARY, font=('Segoe UI', 9),
                                         relief='flat', bd=0,
                                         highlightthickness=1, highlightbackground=BORDER,
                                         highlightcolor=ACCENT)
    settings_output_dir_entry.insert(0, settings.get("bank_output_dir", _default_output_dir()))
    settings_output_dir_entry.pack(side=tk.LEFT, padx=(0, 8), pady=5, fill=tk.X, expand=True)

    def browse_output_dir():
        chosen = filedialog.askdirectory(initialdir=os.path.expanduser("~"), title="Select Output Folder")
        if chosen:
            settings_output_dir_entry.delete(0, tk.END)
            settings_output_dir_entry.insert(0, chosen)
            _sync_settings_from_ui()

    browse_btn = create_styled_button(output_frame, "Browse Folder", browse_output_dir, width=15, accent=False)
    browse_btn.pack(side=tk.LEFT)

    save_btn = create_styled_button(frame, "Save Settings", _sync_settings_from_ui, width=20, accent=True)
    save_btn.pack(pady=14)

    back_button = create_styled_button(frame, "← Back to Menu", initialize_main_menu, width=20)
    back_button.pack(pady=10)
    ensure_window_fits()

def initialize_round_numbers():
    for widget in frame.winfo_children():
        widget.destroy()

    title_label = create_styled_label(frame, "Round Numbers", size=14, bold=True)
    title_label.pack(pady=(10, 10))

    container = tk.Frame(frame, bg=BG_DARK)
    container.pack(fill=tk.BOTH, expand=True, padx=10)

    input_label = create_styled_label(container, "Paste numbers (one per line):", size=9, color=TEXT_SECONDARY)
    input_label.pack(anchor='w')

    input_text = tk.Text(container, height=6, bg=BG_INPUT, fg=TEXT_PRIMARY,
                         insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                         relief='flat')
    input_text.pack(fill=tk.BOTH, expand=True, pady=(3, 10))
    input_resize = attach_auto_resize_text(input_text, min_lines=6, max_lines=14)

    output_label = create_styled_label(container, "Rounded output:", size=9, color=TEXT_SECONDARY)
    output_label.pack(anchor='w')

    output_text = tk.Text(container, height=6, bg=BG_INPUT, fg=TEXT_PRIMARY,
                          insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                          relief='flat')
    output_text.pack(fill=tk.BOTH, expand=True, pady=(3, 10))
    output_resize = attach_auto_resize_text(output_text, min_lines=6, max_lines=14)

    status_label = create_styled_label(container, "", size=9, color=TEXT_SECONDARY)
    status_label.pack(pady=(0, 8))

    def round_numbers():
        raw = input_text.get("1.0", tk.END)
        lines = [line.strip() for line in raw.splitlines() if line.strip()]
        rounded = []
        bad = 0
        for line in lines:
            normalized = line.replace(" ", "").replace(",", ".")
            try:
                value = Decimal(normalized)
                rounded_value = value.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                rounded.append(str(int(rounded_value)))
            except (InvalidOperation, ValueError):
                bad += 1
        output_text.delete("1.0", tk.END)
        output_text.insert(tk.END, "\n".join(rounded))
        output_resize()
        if bad:
            status_label.config(text=f"Skipped {bad} invalid line(s).", fg=TEXT_WARNING)
        else:
            status_label.config(text="Rounded successfully.", fg=TEXT_SUCCESS)

    def copy_output():
        text = output_text.get("1.0", tk.END).strip()
        if text:
            pyperclip.copy(text)
            status_label.config(text="Copied to clipboard (one per row).", fg=TEXT_SUCCESS)
        else:
            status_label.config(text="No output to copy.", fg=TEXT_WARNING)

    btn_row = tk.Frame(container, bg=BG_DARK)
    btn_row.pack(pady=5)

    round_btn = create_styled_button(btn_row, "Round", round_numbers, width=12)
    round_btn.pack(side=tk.LEFT, padx=5)

    copy_btn = create_styled_button(btn_row, "Copy Output", copy_output, width=12, accent=False)
    copy_btn.pack(side=tk.LEFT, padx=5)

    back_button = create_styled_button(frame, "← Back to Menu", initialize_main_menu, width=20)
    back_button.pack(pady=10)
    ensure_window_fits()

def initialize_format_dates():
    for widget in frame.winfo_children():
        widget.destroy()

    title_label = create_styled_label(frame, "Format Dates", size=14, bold=True)
    title_label.pack(pady=(10, 10))

    container = tk.Frame(frame, bg=BG_DARK)
    container.pack(fill=tk.BOTH, expand=True, padx=10)

    input_label = create_styled_label(container, "Paste dates (one per line):", size=9, color=TEXT_SECONDARY)
    input_label.pack(anchor='w')

    input_text = tk.Text(container, height=6, bg=BG_INPUT, fg=TEXT_PRIMARY,
                         insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                         relief='flat')
    input_text.pack(fill=tk.BOTH, expand=True, pady=(3, 10))
    input_resize = attach_auto_resize_text(input_text, min_lines=6, max_lines=14)

    output_label = create_styled_label(container, "Formatted output (DD.MM.YYYY):", size=9, color=TEXT_SECONDARY)
    output_label.pack(anchor='w')

    output_text = tk.Text(container, height=6, bg=BG_INPUT, fg=TEXT_PRIMARY,
                          insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                          relief='flat')
    output_text.pack(fill=tk.BOTH, expand=True, pady=(3, 10))
    output_resize = attach_auto_resize_text(output_text, min_lines=6, max_lines=14)

    status_label = create_styled_label(container, "", size=9, color=TEXT_SECONDARY)
    status_label.pack(pady=(0, 8))

    month_map = {
        "jan": "01", "feb": "02", "mar": "03", "apr": "04",
        "maí": "05", "jun": "06", "jul": "07", "agu": "08",
        "sep": "09", "okt": "10", "nóv": "11", "des": "12",
        "mai": "05", "jún": "06", "júl": "07", "ágú": "08",
        "águ": "08", "nov": "11"
    }

    def format_dates():
        raw = input_text.get("1.0", tk.END)
        lines = [line.strip() for line in raw.splitlines() if line.strip()]
        formatted = []
        bad = 0
        for line in lines:
            cleaned = line.lower().replace("\t", " ").replace(",", " ")
            parts = [p for p in cleaned.split() if p]
            if len(parts) < 3:
                bad += 1
                continue
            day_part = parts[0].replace(".", "")
            month_key = parts[1].strip(".")
            year_part = parts[2]
            if not day_part.isdigit() or not year_part.isdigit():
                bad += 1
                continue
            month_num = month_map.get(month_key)
            if not month_num:
                bad += 1
                continue
            day_num = int(day_part)
            year_num = int(year_part)
            if day_num <= 0 or day_num > 31 or year_num < 1:
                bad += 1
                continue
            formatted.append(f"{day_num:02d}.{month_num}.{year_num:04d}")

        output_text.delete("1.0", tk.END)
        output_text.insert(tk.END, "\n".join(formatted))
        output_resize()
        if bad:
            status_label.config(text=f"Skipped {bad} invalid line(s).", fg=TEXT_WARNING)
        else:
            status_label.config(text="Formatted successfully.", fg=TEXT_SUCCESS)

    def copy_output():
        text = output_text.get("1.0", tk.END).strip()
        if text:
            pyperclip.copy(text)
            status_label.config(text="Copied to clipboard (one per row).", fg=TEXT_SUCCESS)
        else:
            status_label.config(text="No output to copy.", fg=TEXT_WARNING)

    btn_row = tk.Frame(container, bg=BG_DARK)
    btn_row.pack(pady=5)

    format_btn = create_styled_button(btn_row, "Format", format_dates, width=12)
    format_btn.pack(side=tk.LEFT, padx=5)

    copy_btn = create_styled_button(btn_row, "Copy Output", copy_output, width=12, accent=False)
    copy_btn.pack(side=tk.LEFT, padx=5)

    back_button = create_styled_button(frame, "← Back to Menu", initialize_main_menu, width=20)
    back_button.pack(pady=10)
    ensure_window_fits()

def initialize_format_ids():
    for widget in frame.winfo_children():
        widget.destroy()

    title_label = create_styled_label(frame, "Format ID Numbers", size=14, bold=True)
    title_label.pack(pady=(10, 10))

    container = tk.Frame(frame, bg=BG_DARK)
    container.pack(fill=tk.BOTH, expand=True, padx=10)

    input_label = create_styled_label(container, "Paste IDs (one per line):", size=9, color=TEXT_SECONDARY)
    input_label.pack(anchor='w')

    input_text = tk.Text(container, height=6, bg=BG_INPUT, fg=TEXT_PRIMARY,
                         insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                         relief='flat')
    input_text.pack(fill=tk.BOTH, expand=True, pady=(3, 10))
    input_resize = attach_auto_resize_text(input_text, min_lines=6, max_lines=14)

    toggle_frame = tk.Frame(container, bg=BG_DARK)
    toggle_frame.pack(pady=(0, 8))

    mode_var = tk.StringVar(value="remove")
    tk.Radiobutton(toggle_frame, text="Remove dash", variable=mode_var, value="remove",
                   bg=BG_DARK, fg=TEXT_PRIMARY, selectcolor=BG_INPUT,
                   activebackground=BG_DARK, activeforeground=TEXT_PRIMARY).pack(side=tk.LEFT, padx=10)
    tk.Radiobutton(toggle_frame, text="Add dash", variable=mode_var, value="add",
                   bg=BG_DARK, fg=TEXT_PRIMARY, selectcolor=BG_INPUT,
                   activebackground=BG_DARK, activeforeground=TEXT_PRIMARY).pack(side=tk.LEFT, padx=10)

    output_label = create_styled_label(container, "Formatted output:", size=9, color=TEXT_SECONDARY)
    output_label.pack(anchor='w')

    output_text = tk.Text(container, height=6, bg=BG_INPUT, fg=TEXT_PRIMARY,
                          insertbackground=TEXT_PRIMARY, font=('Segoe UI', 10),
                          relief='flat')
    output_text.pack(fill=tk.BOTH, expand=True, pady=(3, 10))
    output_resize = attach_auto_resize_text(output_text, min_lines=6, max_lines=14)

    status_label = create_styled_label(container, "", size=9, color=TEXT_SECONDARY)
    status_label.pack(pady=(0, 8))

    def format_ids():
        raw = input_text.get("1.0", tk.END)
        lines = [line.strip() for line in raw.splitlines() if line.strip()]
        formatted = []
        bad = 0
        for line in lines:
            digits = "".join(ch for ch in line if ch.isdigit())
            if len(digits) < 10:
                bad += 1
                continue
            base = digits[:10]
            if mode_var.get() == "add":
                formatted.append(f"{base[:6]}-{base[6:]}")
            else:
                formatted.append(base)

        output_text.delete("1.0", tk.END)
        output_text.insert(tk.END, "\n".join(formatted))
        output_resize()
        if bad:
            status_label.config(text=f"Skipped {bad} invalid line(s).", fg=TEXT_WARNING)
        else:
            status_label.config(text="Formatted successfully.", fg=TEXT_SUCCESS)

    def copy_output():
        text = output_text.get("1.0", tk.END).strip()
        if text:
            pyperclip.copy(text)
            status_label.config(text="Copied to clipboard (one per row).", fg=TEXT_SUCCESS)
        else:
            status_label.config(text="No output to copy.", fg=TEXT_WARNING)

    btn_row = tk.Frame(container, bg=BG_DARK)
    btn_row.pack(pady=5)

    format_btn = create_styled_button(btn_row, "Format", format_ids, width=12)
    format_btn.pack(side=tk.LEFT, padx=5)

    copy_btn = create_styled_button(btn_row, "Copy Output", copy_output, width=12, accent=False)
    copy_btn.pack(side=tk.LEFT, padx=5)

    back_button = create_styled_button(frame, "← Back to Menu", initialize_main_menu, width=20)
    back_button.pack(pady=10)
    ensure_window_fits()

# =============================================================================
# BANK FORMATTER - Auto-detects Islandsbanki, Landsbanki, or Arion Banki
# =============================================================================

# Bank column configurations: (date_col, text_col, id_col, amount_col, sheet_name_or_None)
BANK_CONFIGS = {
    'islandsbanki': {
        'name': 'Íslandsbanki',
        'columns': ['Dags.', 'Mótaðili', 'Kennitala móttakanda', 'Upphæð'],
        'rename': {'Dags.': 'DATE', 'Mótaðili': 'TEXT', 'Kennitala móttakanda': 'ID', 'Upphæð': 'AMOUNT'},
        'sheet': 'Yfirlit',
        'detect_cols': {'Dags.', 'Kennitala móttakanda'}  # unique identifiers
    },
    'landsbanki': {
        'name': 'Landsbankinn',
        'columns': ['Dags', 'Texti', 'Tilvísun', 'Upphæð'],
        'rename': {'Dags': 'DATE', 'Texti': 'TEXT', 'Tilvísun': 'ID', 'Upphæð': 'AMOUNT'},
        'sheet': None,
        'detect_cols': {'Dags', 'Texti'}  # unique identifiers
    },
    'arion': {
        'name': 'Arion Banki',
        'columns': ['Dagsetning', 'Skýring', 'Tilvísun', 'Upphæð'],
        'rename': {'Dagsetning': 'DATE', 'Skýring': 'TEXT', 'Tilvísun': 'ID', 'Upphæð': 'AMOUNT'},
        'sheet': None,
        'detect_cols': {'Dagsetning', 'Skýring'}  # unique identifiers
    },
    'islandsbanki_innheimta': {
        'name': 'Íslandsbanki Innheimta',
        'columns': ['Kennitala', 'Greiðandi', 'Eindagi', 'Kröfunúmer', 'Upphæð', 'Fjármagnstekjuskattur', 'Dráttarvextir', 'Greidd upphæð', 'Greiðsludagur', 'Rst. Upphæð'],
        'sheet': None,
        'detect_cols': {'Kennitala', 'Greiðsludagur', 'Fjármagnstekjuskattur', 'Dráttarvextir', 'Rst. Upphæð'},
        'custom_processor': True  # uses special processing logic
    },
    'sala_yfirlit': {
        'name': 'Sala Yfirlit',
        'columns': ['Nafn', 'Kennitala', 'Upphæð með vsk', 'Upphæð án vsk', 'Reikningur nr', 'Dagsetning'],
        'sheet': None,
        'detect_cols': {'Nafn', 'Kennitala', 'Upphæð með vsk', 'Reikningur nr', 'Dagsetning'},
        'custom_processor': True  # uses special processing logic
    }
}

def detect_bank_type(file_path):
    """
    Reads the Excel file and detects which bank format it is based on column names.
    Returns (bank_key, df) or (None, None) if unknown.
    """
    def _find_header_row(path, required_cols, max_rows=30):
        try:
            preview = pd.read_excel(path, header=None, nrows=max_rows)
        except Exception:
            return None
        for idx in range(len(preview)):
            row_values = [v for v in preview.iloc[idx].tolist() if pd.notna(v)]
            if not row_values:
                continue
            row_cols = set(str(v).strip() for v in row_values)
            if required_cols.issubset(row_cols):
                return idx
        return None

    # Try Islandsbanki first (has specific sheet)
    try:
        df = pd.read_excel(file_path, sheet_name='Yfirlit')
        cols = set(df.columns)
        if BANK_CONFIGS['islandsbanki']['detect_cols'].issubset(cols):
            return 'islandsbanki', df
    except Exception:
        pass  # Sheet doesn't exist, try others
    
    # Try reading default sheet for other banks
    try:
        df = pd.read_excel(file_path)
        cols = set(df.columns)

        # Check if first row contains metadata headers instead of data columns
        # Arion has "Heiti" / "IBAN númer" at top, Landsbanki has "Netbanki fyrirtækja"
        first_col = str(df.columns[0]).lower() if len(df.columns) > 0 else ''
        has_metadata_header = any(keyword in first_col for keyword in ['heiti', 'netbanki', 'iban'])
        
        # Also check first cell value for metadata text
        if not has_metadata_header and len(df) > 0:
            first_val = str(df.iloc[0, 0]).lower() if pd.notna(df.iloc[0, 0]) else ''
            has_metadata_header = any(keyword in first_val for keyword in ['heiti', 'netbanki', 'iban', 'færslur'])
        
        if has_metadata_header:
            # Try reading with header on row 3 (0-indexed, Excel row 4)
            df_skip3 = pd.read_excel(file_path, header=3)
            cols_skip3 = set(df_skip3.columns)

            # Check Arion (usually row 4)
            if BANK_CONFIGS['arion']['detect_cols'].issubset(cols_skip3):
                return 'arion', df_skip3
            
            # Try reading with header on row 4 (0-indexed, Excel row 5)
            df_skip4 = pd.read_excel(file_path, header=4)
            cols_skip4 = set(df_skip4.columns)
            
            # Check Landsbanki (usually row 5)
            if BANK_CONFIGS['landsbanki']['detect_cols'].issubset(cols_skip4):
                return 'landsbanki', df_skip4
            
            # Check Arion on row 5 as fallback
            if BANK_CONFIGS['arion']['detect_cols'].issubset(cols_skip4):
                return 'arion', df_skip4

        # Islandsbanki Innheimta sometimes has metadata rows above the header.
        innheimta_header = _find_header_row(file_path, BANK_CONFIGS['islandsbanki_innheimta']['detect_cols'])
        if innheimta_header is not None:
            df_innheimta = pd.read_excel(file_path, header=innheimta_header)
            return 'islandsbanki_innheimta', df_innheimta
        
        # Standard detection (no metadata rows)
        # Check Islandsbanki Innheimta first (more specific columns)
        if BANK_CONFIGS['islandsbanki_innheimta']['detect_cols'].issubset(cols):
            return 'islandsbanki_innheimta', df
        # Sala Yfirlit
        if BANK_CONFIGS['sala_yfirlit']['detect_cols'].issubset(cols):
            return 'sala_yfirlit', df
        
        # Check Landsbanki
        if BANK_CONFIGS['landsbanki']['detect_cols'].issubset(cols):
            return 'landsbanki', df
        
        # Check Arion
        if BANK_CONFIGS['arion']['detect_cols'].issubset(cols):
            return 'arion', df
    except Exception:
        pass
    
    return None, None

def initialize_bank_formatter():
    for widget in frame.winfo_children():
        widget.destroy()
    
    title_label = create_styled_label(frame, "Bank Formatter", size=14, bold=True)
    title_label.pack(pady=(10, 20))
    
    display_bank_formatter_controls()

def display_bank_formatter_controls():
    global bank_input_file_entry, bank_status_label, bank_output_name_entry
    global input_drop

    controls_card = tk.Frame(frame, bg=BG_CARD, highlightthickness=1, highlightbackground=BORDER)
    controls_card.pack(fill=tk.X, padx=20, pady=(0, 10))
    controls_frame = tk.Frame(controls_card, bg=BG_CARD)
    controls_frame.pack(fill=tk.X, padx=16, pady=14)

    # Input file row
    # Using two drop areas stacked: top for input, bottom for output
    drop_frame = tk.Frame(controls_frame, bg=BG_CARD)
    drop_frame.pack(fill=tk.X, pady=8)
    
    # Input drop
    input_label = create_styled_label(drop_frame, "Input (drop file here):", size=9, color=TEXT_SECONDARY)
    input_label.pack(anchor='w')
    input_drop = tk.Label(drop_frame, text="Drop input file", bg=BG_INPUT, fg=TEXT_SECONDARY,
                          font=('Segoe UI', 10), width=46, height=4, relief='flat',
                          highlightthickness=1, highlightbackground=BORDER)
    input_drop.pack(fill=tk.X, pady=(4, 8))
    
    bank_input_file_entry = create_styled_entry(drop_frame, width=35)
    bank_input_file_entry.pack_forget()
    
    def on_drop_input(path):
        if path:
            bank_input_file_entry.delete(0, tk.END)
            bank_input_file_entry.insert(0, path)
            input_drop.config(text=os.path.basename(path), fg=TEXT_PRIMARY)
            # Auto-detect and update status
            bank_type, _ = detect_bank_type(path)
            if bank_type:
                bank_status_label.config(text=f"✓ Detected: {BANK_CONFIGS[bank_type]['name']}", fg=TEXT_SUCCESS)
            else:
                bank_status_label.config(text="✗ Unknown bank format", fg=TEXT_ERROR)
    if not attach_drop_target(input_drop, on_drop_input):
        input_drop.config(text="Install tkinterdnd2 for drag-and-drop", fg=TEXT_WARNING)
    
    browse_input_btn = create_styled_button(drop_frame, "Browse Input", browse_bank_input_file, width=15, accent=False)
    browse_input_btn.pack(pady=(2, 12))

    # Output name (optional)
    output_name_label = create_styled_label(drop_frame, "Output file name (optional):", size=9, color=TEXT_SECONDARY)
    output_name_label.pack(anchor='w')
    bank_output_name_entry = create_styled_entry(drop_frame, width=35)
    bank_output_name_entry.pack(fill=tk.X, pady=(3, 10))

    output_info = create_styled_label(drop_frame, f"Output folder: {settings.get('bank_output_dir', _default_output_dir())}", size=9, color=TEXT_SECONDARY)
    output_info.pack(pady=(6, 10))

    # Status label
    bank_status_label = create_styled_label(controls_frame, "Select input file to auto-detect bank", size=9, color=TEXT_SECONDARY)
    bank_status_label.pack(pady=(6, 16))
    
    # Buttons
    run_btn = create_styled_button(controls_frame, "▶ Run", lambda: run_bank_formatter_script(autofill_7810=True), width=20)
    run_btn.pack(pady=(6, 10))

    run_no_fill_btn = create_styled_button(controls_frame, "Run without 7810", lambda: run_bank_formatter_script(autofill_7810=False), width=20, accent=False)
    run_no_fill_btn.configure(bg=BG_INPUT, activebackground=BG_INPUT, highlightbackground=BORDER)
    run_no_fill_btn.pack(pady=(0, 10))
    
    back_btn = create_styled_button(controls_frame, "← Back to Menu", initialize_main_menu, width=20, accent=False)
    back_btn.pack(pady=(0, 4))
    ensure_window_fits()

def browse_bank_input_file():
    filename = filedialog.askopenfilename(
        initialdir=os.getcwd(), 
        title="Select Input File", 
        filetypes=(("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*"))
    )
    if filename:
        bank_input_file_entry.delete(0, tk.END)
        bank_input_file_entry.insert(0, filename)
        input_drop.config(text=os.path.basename(filename), fg=TEXT_PRIMARY)
        # Auto-detect bank type and update status
        bank_type, _ = detect_bank_type(filename)
        if bank_type:
            bank_name = BANK_CONFIGS[bank_type]['name']
            bank_status_label.config(text=f"✓ Detected: {bank_name}", fg=TEXT_SUCCESS)
        else:
            bank_status_label.config(text="✗ Unknown bank format", fg=TEXT_ERROR)

def _bank_output_path(output_name):
    output_dir = settings.get("bank_output_dir") or _default_output_dir()
    os.makedirs(output_dir, exist_ok=True)
    output_name = output_name.strip()
    if output_name:
        if not output_name.lower().endswith(".xlsx"):
            output_name += ".xlsx"
        filename = output_name
    else:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"innlestur_{timestamp}.xlsx"
    return os.path.join(output_dir, filename)

def run_bank_formatter_script(autofill_7810=True):
    input_file_path = bank_input_file_entry.get()
    output_name = bank_output_name_entry.get() if "bank_output_name_entry" in globals() else ""
    output_file_path = _bank_output_path(output_name)
    
    if not input_file_path:
        bank_status_label.config(text="⚠ Please select an input file", fg=TEXT_WARNING)
        return
    
    bank_type, df = detect_bank_type(input_file_path)
    
    if bank_type is None:
        bank_status_label.config(text="✗ Could not detect bank format", fg=TEXT_ERROR)
        play_error_sound()
        return
    
    config = BANK_CONFIGS[bank_type]
    
    try:
        # Check if this bank needs custom processing
        if config.get('custom_processor'):
            if bank_type == 'islandsbanki_innheimta':
                process_islandsbanki_innheimta(df, output_file_path, config['name'], autofill_7810=autofill_7810)
            elif bank_type == 'sala_yfirlit':
                process_sala_yfirlit(df, output_file_path, config['name'], autofill_7810=autofill_7810)
        else:
            # Standard processing for regular banks
            extracted_columns = df[config['columns']].copy()
            extracted_columns['Positive/Negative'] = extracted_columns[config['columns'][3]].apply(lambda x: '+' if x >= 0 else '-')
            extracted_columns[config['columns'][3]] = extracted_columns[config['columns'][3]].abs()
            
            # Rename to standard format
            extracted_columns.rename(columns=config['rename'], inplace=True)
            
            # Insert DEBIT and CREDIT columns
            extracted_columns.insert(2, 'DEBIT', '')
            extracted_columns.insert(5, 'CREDIT', '')

            if autofill_7810:
                extracted_columns.loc[extracted_columns['Positive/Negative'] == '+', 'DEBIT'] = 7810
                extracted_columns.loc[extracted_columns['Positive/Negative'] == '-', 'CREDIT'] = 7810
            
            # Sort by Positive/Negative and TEXT
            extracted_columns = extracted_columns.sort_values(by=['Positive/Negative', 'TEXT'])
            
            # Convert DATE to text to preserve format on copy-paste
            extracted_columns['DATE'] = extracted_columns['DATE'].apply(format_date_as_text)
            
            # Save to Excel
            extracted_columns.to_excel(output_file_path, index=False)
            
            # Auto-fit column widths
            autofit_excel_columns(output_file_path)
            
            display_bank_formatter_success(config['name'], output_file_path)
        
    except Exception as e:
        bank_status_label.config(text=f"✗ Error: {str(e)[:40]}", fg=TEXT_ERROR)
        play_error_sound()

def format_date_as_text(date_val):
    """Convert date value to text string format (DD.MM.YYYY)."""
    if pd.isna(date_val):
        return ''
    if isinstance(date_val, pd.Timestamp):
        return date_val.strftime('%d.%m.%Y')
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime('%d.%m.%Y')
    if isinstance(date_val, datetime.date):
        return date_val.strftime('%d.%m.%Y')
    # Already a string or other type
    return str(date_val)

def autofit_excel_columns(file_path):
    """Auto-fit column widths in Excel file and format DATE column as text."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Find DATE column index
    date_col_idx = None
    for idx, cell in enumerate(ws[1], 1):  # First row (headers)
        if cell.value == 'DATE':
            date_col_idx = idx
            break
    
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            # Set DATE column format to text (@) to preserve text format on copy-paste
            if date_col_idx and col[0].column == date_col_idx:
                cell.number_format = '@'
        ws.column_dimensions[column].width = max_length + 2
    wb.save(file_path)

def _round_half_up_decimal(val):
    try:
        return Decimal(str(val).replace(" ", "").replace(",", "."))
    except Exception:
        return None

def _parse_innheimta_amount(val, force_thousands=False):
    if pd.isna(val):
        return 0
    if isinstance(val, (int,)):
        if force_thousands and abs(val) < 1000:
            return int(val) * 1000
        return int(val)
    if isinstance(val, (float, Decimal)):
        try:
            dec = Decimal(str(val))
        except Exception:
            return 0
        if dec == dec.to_integral_value():
            num = int(dec)
            if force_thousands and abs(num) < 1000:
                return num * 1000
            return num
        # No decimals in source files; non-integer means thousands separator was read as decimal.
        return int((dec * 1000).to_integral_value(rounding=ROUND_HALF_UP))
    text = str(val).strip()
    if not text:
        return 0
    text = text.replace(" ", "").replace("\u00a0", "")
    sign = 1
    if text.startswith("(") and text.endswith(")"):
        sign = -1
        text = text[1:-1]
    if text.startswith("-"):
        sign = -1
        text = text[1:]
    elif text.startswith("+"):
        text = text[1:]
    if not text:
        return 0
    # No decimals in these files; dots are thousands separators.
    text = text.replace(".", "")
    text = text.replace(",", "")
    if not text:
        return 0
    try:
        return sign * int(text)
    except Exception:
        try:
            return sign * int(Decimal(text))
        except Exception:
            return 0


def _format_kennitala(val):
    if pd.isna(val):
        return ''
    text = str(val).strip()
    if not text:
        return ''
    text = text.replace(" ", "").replace("-", "")
    text = text.replace(".", "")
    return text


def process_islandsbanki_innheimta(df, output_file_path, bank_name, autofill_7810=True):
    columns_to_select = ['Kennitala', 'Greiðsludagur', 'Fjármagnstekjuskattur', 'Dráttarvextir', 'Rst. Upphæð', 'Upphæð']
    df_selected = df[columns_to_select].copy()

    entries = []
    for _, row in df_selected.iterrows():
        date_val = format_date_as_text(row['Greiðsludagur'])
        id_val = _format_kennitala(row['Kennitala'])
        text_val = 'kostnaður'
        fjarmagn = _parse_innheimta_amount(row['Fjármagnstekjuskattur'])
        drattar = _parse_innheimta_amount(row['Dráttarvextir'])
        rst_amount = _parse_innheimta_amount(row['Rst. Upphæð'], force_thousands=True)
        upphaed = _parse_innheimta_amount(row['Upphæð'], force_thousands=True)

        if fjarmagn != 0:
            entries.append({
                'DATE': date_val,
                'TEXT': text_val,
                'DEBIT': 7660,
                'AMOUNT': fjarmagn,
                'ID': id_val,
                'CREDIT': 7620
            })
        if drattar != 0:
            entries.append({
                'DATE': date_val,
                'TEXT': text_val,
                'DEBIT': 7620,
                'AMOUNT': drattar,
                'ID': id_val,
                'CREDIT': 6100
            })
        # Rst. Upphæð + Fjármagnstekjuskattur - Dráttarvextir - Upphæð (J + F - G - E).
        calc_amount = rst_amount + fjarmagn - drattar - upphaed
        if calc_amount != 0:
            entries.append({
                'DATE': date_val,
                'TEXT': text_val,
                'DEBIT': 7620,
                'AMOUNT': calc_amount,
                'ID': id_val,
                'CREDIT': 4400
            })

    df_output = pd.DataFrame(entries, columns=['DATE', 'TEXT', 'DEBIT', 'AMOUNT', 'ID', 'CREDIT'])
    df_output.to_excel(output_file_path, index=False)
    autofit_excel_columns(output_file_path)
    display_bank_formatter_success(bank_name, output_file_path)


def _parse_icelandic_date(val):
    if pd.isna(val):
        return ''
    if isinstance(val, (datetime.datetime, datetime.date, pd.Timestamp)):
        return format_date_as_text(val)
    text = str(val).strip()
    month_map = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
        'maí': 5, 'jun': 6, 'jul': 7, 'agu': 8,
        'sep': 9, 'okt': 10, 'nóv': 11, 'des': 12,
        'mai': 5, 'jún': 6, 'júl': 7, 'ágú': 8,
        'águ': 8, 'nov': 11
    }
    parts = text.replace(",", " ").replace("\t", " ").split()
    # Expected like "31. júl 2025"
    if len(parts) >= 3:
        day_raw = parts[0].replace(".", "")
        month_raw = parts[1].strip(".").lower()
        year_raw = parts[2]
        if day_raw.isdigit() and year_raw.isdigit() and month_raw in month_map:
            try:
                dt = datetime.datetime(int(year_raw), month_map[month_raw], int(day_raw))
                return dt.strftime("%d.%m.%Y")
            except Exception:
                pass
    return text  # fallback to original

def process_sala_yfirlit(df, output_file_path, bank_name, autofill_7810=True):
    # Drop total rows if present (where Kennitala is NaN and/or Nafn starts with 'Alls')
    df_clean = df.copy()
    df_clean = df_clean[~df_clean['Nafn'].astype(str).str.lower().str.startswith('alls')]
    cols = ['Nafn', 'Kennitala', 'Upphæð með vsk', 'Reikningur nr', 'Dagsetning']
    df_selected = df_clean[cols].copy()

    # ID: strip dash and spaces
    df_selected['ID'] = df_selected['Kennitala'].astype(str).str.replace('-', '').str.replace(' ', '')

    # TEXT: "Reikningur X" (strip trailing .0)
    def _fmt_reikningur(x):
        try:
            as_float = float(x)
            if as_float.is_integer():
                return f"Reikningur {int(as_float)}"
        except Exception:
            pass
        return f"Reikningur {str(x).rstrip('0').rstrip('.') if isinstance(x, str) else x}"
    df_selected['TEXT'] = df_selected['Reikningur nr'].apply(_fmt_reikningur)

    # AMOUNT: round half up from Upphæð með vsk
    def _round_amount(v):
        dec = _round_half_up_decimal(v)
        if dec is None:
            return v
        return int(dec.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    df_selected['AMOUNT'] = df_selected['Upphæð með vsk'].apply(_round_amount)

    # DATE: format as text
    df_selected['DATE'] = df_selected['Dagsetning'].apply(_parse_icelandic_date)

    # Positive/Negative and absolute amount
    df_selected['Positive/Negative'] = df_selected['AMOUNT'].apply(lambda x: '+' if _round_half_up_decimal(x) is None or _round_half_up_decimal(x) >= 0 else '-')
    df_selected['AMOUNT'] = df_selected['AMOUNT'].apply(lambda x: abs(int(_round_half_up_decimal(x).quantize(Decimal("1"), rounding=ROUND_HALF_UP))) if _round_half_up_decimal(x) is not None else x)

    # Insert DEBIT and CREDIT
    df_selected.insert(2, 'DEBIT', '')
    df_selected.insert(5, 'CREDIT', '')

    if autofill_7810:
        df_selected['DEBIT'] = 7620
        df_selected['CREDIT'] = 1000

    df_selected = df_selected.sort_values(by=['Positive/Negative', 'TEXT'])

    df_final = df_selected[['DATE', 'TEXT', 'DEBIT', 'AMOUNT', 'ID', 'CREDIT']].copy()
    df_final.to_excel(output_file_path, index=False)
    autofit_excel_columns(output_file_path)
    display_bank_formatter_success(bank_name, output_file_path)

def display_bank_formatter_success(bank_name, output_file_path):
    for widget in frame.winfo_children():
        widget.destroy()
    play_success_sound()
    
    # Success header
    success_label = create_styled_label(frame, "✓ Success!", size=18, color=TEXT_SUCCESS, bold=True)
    success_label.pack(pady=(30, 15))
    
    # Bank info card
    card = tk.Frame(frame, bg=BG_CARD, padx=30, pady=20)
    card.pack(padx=20, pady=10)
    
    tk.Label(card, text=f"Processed as:", bg=BG_CARD, fg=TEXT_SECONDARY,
             font=('Segoe UI', 10)).pack(pady=3)
    tk.Label(card, text=bank_name, bg=BG_CARD, fg=TEXT_PRIMARY,
             font=('Segoe UI', 12, 'bold')).pack(pady=3)

    output_name = os.path.basename(output_file_path) if output_file_path else ""
    if output_name:
        tk.Label(card, text="Output file:", bg=BG_CARD, fg=TEXT_SECONDARY,
                 font=('Segoe UI', 10)).pack(pady=(12, 3))
        tk.Label(card, text=output_name, bg=BG_CARD, fg=TEXT_PRIMARY,
                 font=('Segoe UI', 10, 'bold')).pack(pady=3)

        open_btn = create_styled_button(frame, "Open Output File", lambda: os.startfile(output_file_path), width=20, accent=False)
        open_btn.pack(pady=(5, 10))
    
    back_button = create_styled_button(frame, "← Back to Menu", initialize_main_menu, width=20)
    back_button.pack(pady=25)
    ensure_window_fits()

def play_success_sound():
    sound_file_path = resource_path(os.path.join("tok", "success.wav"))
    winsound.PlaySound(sound_file_path, winsound.SND_FILENAME | winsound.SND_ASYNC)

def play_error_sound():
    error_sound_file_path = resource_path(os.path.join("tok", "error.wav"))
    winsound.PlaySound(error_sound_file_path, winsound.SND_FILENAME | winsound.SND_ASYNC)

def on_watchdog_timeout():
    logging.error("Watchdog timer exceeded 15 seconds. Playing error sound.")
    play_error_sound()

# =============================================================================
# MAIN APPLICATION SETUP
# =============================================================================

if TKDND_AVAILABLE:
    root = TkinterDnD.Tk()
else:
    root = tk.Tk()
settings = load_settings()
action_delay = settings.get("action_delay", 0.1)
start_delay = settings.get("start_delay", 3)
root.title("Tok Tenging")
root.geometry("520x640")
root.configure(bg=BG_DARK)
root.resizable(True, True)
root.minsize(480, 600)

# Configure ttk styles (progressbar only)
style = ttk.Style()
style.theme_use('clam')
style.configure('Horizontal.TProgressbar',
                background=ACCENT,
                troughcolor=BG_INPUT,
                bordercolor=BG_DARK,
                lightcolor=ACCENT,
                darkcolor=ACCENT)

# Header with smaller logo
header_frame = tk.Frame(root, bg=BG_DARK)
header_frame.pack(fill=tk.X, pady=(10, 0))

logo_path = resource_path(os.path.join("tok", "temp2.png"))
logo_image = PhotoImage(file=logo_path)
logo_image_small = logo_image.subsample(2, 2)
logo_label = tk.Label(header_frame, image=logo_image_small, bg=BG_DARK)
logo_label.image = logo_image_small
logo_label.pack()
root.iconphoto(True, logo_image)

# Main content frame - simple, no scroll needed with proper sizing
frame = tk.Frame(root, bg=BG_DARK)
frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)

initialize_main_menu()

root.mainloop()
